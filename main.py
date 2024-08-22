from openpyxl import load_workbook
from Magic import Maker
import time

# Загрузите реестр
wb = load_workbook('test.xlsx')
sheet = wb.active
num = 0
last_id = 0
t = time.time()
try:
    for row in sheet.iter_rows():
        num += 1
        co = 0
        try:
            for cell in row:
                prob = False
                if cell.coordinate[-1] == '1':
                    continue

                co += 1
                match co:
                    case 5:
                        check_num = cell.value.strip()
                    case 15:
                        meter = cell.value.strip()
                    case 16:
                        id = cell.value.strip()
                    case 19:
                        data = cell.value.strip()
                    case 22:
                        date = cell.value
                        n = 1
                        while n != 6:
                            try:
                                M = Maker((meter, id, data, date), num, check_num)
                                new_path = M.get()
                                if new_path:
                                    break
                            except Exception as e:
                                print(f"Error with {check_num}, attempt {n}, {e}")
                            n += 1
                        prob = True
                    case 30:
                        if not prob:
                            cell.value = str(new_path)
                            wb.save('data.xlsx')
                            last_id = check_num
        except Exception as e:
            print(f"Error {e}", last_id)
            continue
except Exception as e:
    print(f"Error {e}, last file {last_id}")
print(time.time() - t)